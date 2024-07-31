import requests
from bs4 import BeautifulSoup
import openpyxl

#excel
workbook=openpyxl.Workbook()
hoja_activa=workbook.active

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36 Brave/91.0.4472.124'
}

url=requests.get('https://www.infotec.com.pe/3-laptops-y-notebooks')

soup=BeautifulSoup(url.content, 'html.parser')
################################################################################################
nombres=[]

elementos_a=soup.find_all('h2', class_='h3 product-title')

for elemento_h2 in elementos_a:
    elemento_a=elemento_h2.find('a')
    if elemento_a:
        nombres.append(elemento_a.text.strip())
    
# print(nombres)
################################################################################################
precios=[]

elementos_p=soup.find_all('span', class_='product-price')

for elemento_p in elementos_p:
    precios.append(elemento_p.text.strip())
    
# print(precios)
################################################################################################

for i in range(len(nombres)):
    hoja_activa.cell(row=i+1, column=1, value=nombres[i])
    hoja_activa.cell(row=i+1, column=2, value=precios[i])
    
workbook.save("scraping.xlsx")
    